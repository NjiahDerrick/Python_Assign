import pandas as pd
import csv

from os import P_DETACH
from openpyxl import Workbook, load_workbook

wb = load_workbook('Employeedata.xlsx')

ws = wb.active


range = ws["B2":"B30"]

def employee_data_base():
  for cell in range:
    for x in cell:
      print(x.value)
employee_data_base()
print("****************************************Below is the updated Database********************************************")

def updated_employee_database():
  for cell in range:
    for x in cell:
      text = x.value
      changeSufix = text.replace("helpinghands.cm","handsinhands.org")
      x.value = changeSufix
      print(changeSufix)
      wb.save('modifiedemployeedata.xlsx')
      

updated_employee_database()

print("*********************************************CSV*******************************************************************")



df = pd.read_csv("Employeedata.csv")

print(df)

print("***********************************UPDATED CSV FILE**********************************************")


df['Email Addresses'] = df['Email Addresses'].str.replace('helpinghands.cm', 'handsinhands.org')
print(df)

df.to_csv("modifiedEmployee.csv", index=False)